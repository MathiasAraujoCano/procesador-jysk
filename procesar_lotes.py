import pandas as pd
import tkinter as tk
from tkinter import filedialog
import os
from openpyxl.styles import PatternFill, Font

def main():
    # Configurar tkinter para mejor compatibilidad con Windows
    root = tk.Tk()
    root.withdraw()
    
    # En Windows, es mejor mostrar la ventana temporalmente para evitar problemas
    if os.name == 'nt':  # Windows
        root.attributes('-topmost', True)
        root.update()

    print("Selecciona el archivo de Descarga.xlsx")
    file_descarga = filedialog.askopenfilename(
        title="Selecciona el archivo en formato xlsx", 
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
    )
    
    if not file_descarga:
        print("❌ No se seleccionó ningún archivo.")
        return

    # 🗂️ MAPEO DE CÓDIGOS DE CAJA A NOMBRES DE CAJA (hardcodeado en el script)
    caja_nombre_map = {
        'T00003': '1-CO',
        'T00004': '2-CO', 
        'T00005': '3-CO',
        'T00006': '4-CO',
        'T00007': '5-CO',
        'T00001': '6-CO Manual',
        'T00002': '7-CO Manual',
        # Agregar más códigos si aparecen
    }
    
    print("🗂️ Mapeo de códigos de caja configurado:")
    for codigo, nombre in caja_nombre_map.items():
        print(f"  Código {codigo} → Caja {nombre}")

    # Leer archivo
    try:
        df = pd.read_excel(file_descarga)
        print(f"📊 Archivo leído: {len(df)} registros")
    except Exception as e:
        print(f"❌ Error leyendo el archivo: {e}")
        return

    # 🧹 LIMPIEZA SEGÚN ESPECIFICACIONES
    print(f"📊 Registros antes de limpieza: {len(df)}")
    
    # 1. Eliminar Denegadas y Reversadas
    df_original = len(df)
    df = df[~df['Estado'].isin(['Denegada', 'Reversada'])]
    eliminadas_estado = df_original - len(df)
    print(f"🗑️  Eliminadas por estado (Denegada/Reversada): {eliminadas_estado}")
    
    # 2. Manejar Anuladas: Eliminar tanto la anulada como la aprobada con mismo nro de autorización
    if "Autorización" in df.columns and "Operación" in df.columns:
        df_antes_anuladas = len(df)
        # Buscar anulaciones en la columna 'Operación' (no 'Estado')
        anuladas = df[df['Operación'] == 'Anulación']['Autorización'].unique()
        print(f"🔍 Encontradas {len(anuladas)} autorizaciones con anulación")
        # Eliminar todas las transacciones (Anulación y su par Aprobada) que tengan el mismo número de autorización
        df = df[~df['Autorización'].isin(anuladas)]
        eliminadas_anuladas = df_antes_anuladas - len(df)
        print(f"🗑️  Eliminadas por anulación (incluye pares): {eliminadas_anuladas}")
    
    # 3. Las "Devolución" se mantienen (no se eliminan)
    devoluciones = len(df[df['Estado'] == 'Devolución'])
    if devoluciones > 0:
        print(f"↩️  Devoluciones mantenidas: {devoluciones}")
    
    # 4. Normalizar nombres de tarjetas: Maestro → Mastercard
    df['Tarjeta'] = df['Tarjeta'].replace({'Maestro': 'Mastercard'})
    
    print(f"📊 Registros después de limpieza: {len(df)}")
    print(f"📊 Estados finales: {', '.join(sorted(df['Estado'].unique()))}")
    print(f"💳 Tarjetas: {', '.join(sorted(df['Tarjeta'].unique()))}")

    # 🔧 El archivo ya tiene una columna "Caja" con códigos como T00003
    # Vamos a crear una nueva columna "Caja_Nombre" con los nombres legibles
    
    print("🔍 Códigos de caja encontrados en el archivo:")
    codigos_caja_unicos = df['Caja'].unique()
    
    # Función para mapear código de caja a nombre
    def mapear_nombre_caja(codigo_caja):
        """Convierte el código de caja (T00003) al nombre legible (1-CO)"""
        codigo_str = str(codigo_caja).strip()
        return caja_nombre_map.get(codigo_str, f"Sin mapear: {codigo_str}")
    
    # Aplicar el mapeo para crear la columna con nombres legibles
    df['Caja_Nombre'] = df['Caja'].apply(mapear_nombre_caja)
    
    # Mostrar el mapeo encontrado
    for codigo in sorted(codigos_caja_unicos):
        if pd.notna(codigo):
            nombre = mapear_nombre_caja(codigo)
            estado = "✅" if codigo in caja_nombre_map else "❌"
            print(f"  {estado} Código: {codigo} → Nombre: {nombre}")
    
    # 🔧 Verificar si el merge fue exitoso
    if 'Caja' not in df.columns:
        print("❌ Error: No se encontró la columna 'Caja' después del merge.")
        print("Verificar que el archivo Tablacajasxsuc.xlsx tenga las columnas correctas.")
        return
    
    # 🔧 Verificar si hay cajas con mapeo
    cajas_con_datos = df[df['Caja_Nombre'].str.contains('Sin mapear') == False]
    if cajas_con_datos.empty:
        print("❌ Error: No se encontraron códigos de caja con mapeo válido.")
        print("Códigos en archivo:", sorted([str(c) for c in df['Caja'].unique() if pd.notna(c)]))
        print("Mapeo disponible:", list(caja_nombre_map.keys()))
        return
    
    # 🔧 Mostrar información del merge
    total_registros = len(df)
    registros_con_caja = len(cajas_con_datos)
    print(f"📊 Registros totales: {total_registros}")
    print(f"📊 Registros con caja asignada: {registros_con_caja}")
    
    if registros_con_caja < total_registros:
        sin_mapeo = df[df['Caja_Nombre'].str.contains('Sin mapear')]
        print(f"⚠️  Registros sin mapeo de caja: {len(sin_mapeo)}")
        print("Códigos sin mapeo:", sorted([str(c) for c in sin_mapeo['Caja'].unique() if pd.notna(c)]))

    # � Aplicar normalización de tarjetas también a los datos filtrados
    cajas_con_datos['Tarjeta'] = cajas_con_datos['Tarjeta'].replace({'MAESTRO': 'MASTERCARD'})
    
    print(f"💳 Tarjetas después de normalización final: {', '.join(sorted(cajas_con_datos['Tarjeta'].unique()))}")
    
    # �📊 CREAR RESUMEN AGRUPADO
    print("\n📊 Creando resumen agrupado...")
    
    # 💰 AJUSTAR MONTOS: Las devoluciones deben restarse, no sumarse
    def calcular_monto_ajustado(row):
        """Ajusta el monto según el tipo de operación: Devolución se resta"""
        if row['Operación'] == 'Devolución':
            return -abs(row['Monto'])  # Convertir a negativo (restar)
        else:
            return row['Monto']  # Mantener positivo (sumar)
    
    # Aplicar el ajuste de montos antes de agrupar
    cajas_con_datos = cajas_con_datos.copy()  # Evitar warning de modificación
    cajas_con_datos['Monto_Ajustado'] = cajas_con_datos.apply(calcular_monto_ajustado, axis=1)
    
    print(f"💰 Montos ajustados (devoluciones como negativas):")
    devoluciones_count = len(cajas_con_datos[cajas_con_datos['Operación'] == 'Devolución'])
    if devoluciones_count > 0:
        monto_devoluciones = cajas_con_datos[cajas_con_datos['Operación'] == 'Devolución']['Monto_Ajustado'].sum()
        print(f"     📊 {devoluciones_count} devoluciones por ${monto_devoluciones:,.2f}")
    
    # Agrupar por Fecha, Lote, Caja_Nombre y Tarjeta usando el monto ajustado
    resumen = cajas_con_datos.groupby(['Fecha', 'Lote', 'Caja_Nombre', 'Tarjeta']).agg({
        'Monto_Ajustado': ['sum', 'count'],  # Suma con devoluciones negativas y cantidad de transacciones
        'Operación': lambda x: ', '.join(x.unique()),  # Tipos de operación
        'Estado': lambda x: ', '.join(x.unique()),  # Estados únicos
        'Caja': 'first'  # Código de caja original
    }).round(2)
    
    # Aplanar las columnas multi-nivel
    resumen.columns = ['Monto_Total', 'Cantidad_Transacciones', 'Operaciones', 'Estados', 'Codigo_Caja']
    resumen = resumen.reset_index()
    
    # Ordenar por fecha, lote y caja
    resumen = resumen.sort_values(['Fecha', 'Lote', 'Caja_Nombre', 'Tarjeta'])
    
    # CREAR EXCEL SOLO CON HOJAS POR FECHA
    from datetime import datetime
    
    # 🎨 DEFINIR COLORES POR CAJA
    colores_caja = {
        '1-CO': 'FFE6E6',          # Rojo claro
        '2-CO': 'E6F3FF',          # Azul claro
        '3-CO': 'E6FFE6',          # Verde claro
        '4-CO': 'FFF2E6',          # Naranja claro
        '5-CO': 'F0E6FF',          # Morado claro
        '6-CO Manual': 'FFFFE6',   # Amarillo claro
        '7-CO Manual': 'F0F0F0'    # Gris claro
    }
    
    # Generar nombre de archivo con fecha y hora actual
    ahora = datetime.now()
    timestamp = ahora.strftime("%d%m%Y_%H%M")  # formato: ddmmyyyy_hhmm
    
    # Usar separadores de ruta apropiados para el sistema operativo
    output_dir = os.path.dirname(file_descarga)
    output_file = os.path.join(output_dir, f"resumen_{timestamp}.xlsx")
    
    # En Windows, asegurar que el directorio sea accesible
    if os.name == 'nt':  # Windows
        try:
            os.makedirs(output_dir, exist_ok=True)
        except:
            # Si hay problemas con el directorio, usar el escritorio
            output_dir = os.path.join(os.path.expanduser("~"), "Desktop")
            output_file = os.path.join(output_dir, f"resumen_{timestamp}.xlsx")
    
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        
        # 📋 CREAR SOLO HOJAS SEPARADAS POR FECHA
        fechas_unicas = resumen['Fecha'].unique()
        
        for fecha in sorted(fechas_unicas):
            datos_fecha = resumen[resumen['Fecha'] == fecha].copy()
            
            # Ordenar por Caja → Lote → Tarjeta
            datos_fecha = datos_fecha.sort_values(['Caja_Nombre', 'Lote', 'Tarjeta'])
            
            # 🔧 Eliminar columnas no deseadas: Fecha, Operaciones, Estados  
            columnas_mantener = ['Lote', 'Caja_Nombre', 'Tarjeta', 'Monto_Total', 'Cantidad_Transacciones', 'Codigo_Caja']
            datos_fecha_limpio = datos_fecha[columnas_mantener]
            
            # 📊 AGREGAR SUBTOTALES POR LOTE
            datos_con_subtotales = []
            
            # Agrupar por Caja y Lote para insertar subtotales
            for (caja, lote), grupo_lote in datos_fecha_limpio.groupby(['Caja_Nombre', 'Lote']):
                # Agregar las filas del lote (convertir Series a dict para consistencia)
                for _, fila in grupo_lote.iterrows():
                    datos_con_subtotales.append(fila.to_dict())
                
                # Calcular subtotal del lote
                subtotal_monto = grupo_lote['Monto_Total'].sum()
                subtotal_transacciones = grupo_lote['Cantidad_Transacciones'].sum()
                
                # Crear fila de subtotal
                fila_subtotal = {
                    'Lote': lote,
                    'Caja_Nombre': caja,
                    'Tarjeta': f'SUBTOTAL LOTE {lote}',
                    'Monto_Total': float(subtotal_monto),  # Asegurar que sea float
                    'Cantidad_Transacciones': int(subtotal_transacciones),  # Asegurar que sea int
                    'Codigo_Caja': grupo_lote['Codigo_Caja'].iloc[0]
                }
                datos_con_subtotales.append(fila_subtotal)
                
                # Agregar fila vacía para separar visualmente
                fila_vacia = {
                    'Lote': '',
                    'Caja_Nombre': '',
                    'Tarjeta': '',
                    'Monto_Total': '',
                    'Cantidad_Transacciones': '',
                    'Codigo_Caja': ''
                }
                datos_con_subtotales.append(fila_vacia)
            
            # Convertir a DataFrame
            if datos_con_subtotales:  # Solo si hay datos
                datos_finales = pd.DataFrame(datos_con_subtotales)
                
                # 🔧 Asegurar tipos de datos correctos para evitar problemas en Excel
                datos_finales['Monto_Total'] = pd.to_numeric(datos_finales['Monto_Total'], errors='coerce')
                datos_finales['Cantidad_Transacciones'] = pd.to_numeric(datos_finales['Cantidad_Transacciones'], errors='coerce')
            else:
                # Si no hay datos, crear DataFrame vacío con las columnas correctas
                datos_finales = pd.DataFrame(columns=['Lote', 'Caja_Nombre', 'Tarjeta', 'Monto_Total', 'Cantidad_Transacciones', 'Codigo_Caja'])
            
            # Crear nombre de hoja limpio y simple
            try:
                # Si fecha es un objeto datetime, formatear directamente
                if hasattr(fecha, 'strftime'):
                    fecha_str = fecha.strftime("%d-%m-%Y")
                else:
                    # Si es string, intentar convertir y formatear
                    from datetime import datetime
                    if isinstance(fecha, str):
                        # Intentar diferentes formatos comunes
                        for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d"]:
                            try:
                                fecha_obj = datetime.strptime(fecha, fmt)
                                fecha_str = fecha_obj.strftime("%d-%m-%Y")
                                break
                            except:
                                continue
                        else:
                            # Si no se puede parsear, usar string limpio
                            fecha_str = str(fecha).replace('/', '-').replace(' 00:00:00', '')[:10]
                    else:
                        fecha_str = str(fecha).replace('/', '-').replace(' 00:00:00', '')[:10]
                
                sheet_name = f"Fecha {fecha_str}"
                
            except Exception:
                # Fallback: usar el método anterior pero mas limpio
                fecha_str = str(fecha).replace('/', '-').replace(' 00:00:00', '').replace(' ', '')[:10]
                sheet_name = f"Fecha {fecha_str}"
            
            # Guardar en hoja separada solo si tiene datos
            if not datos_finales.empty:
                datos_finales.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # 🎨 APLICAR FORMATO Y COLORES
                worksheet = writer.sheets[sheet_name]
                
                # Estilos
                font_negrita = Font(bold=True)
                
                # Iterar por las filas para aplicar formato
                for row_idx, (_, row) in enumerate(datos_finales.iterrows(), start=2):  # start=2 porque row 1 son headers
                    caja_nombre = row['Caja_Nombre']
                    tarjeta = row['Tarjeta']
                    
                    # Aplicar color de fondo según la caja
                    if caja_nombre in colores_caja:
                        color_hex = colores_caja[caja_nombre]
                        fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
                        
                        # Aplicar color a todas las celdas de la fila
                        for col_idx in range(1, len(datos_finales.columns) + 1):
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            cell.fill = fill
                    
                    # Aplicar negrita a los subtotales
                    if isinstance(tarjeta, str) and tarjeta.startswith('SUBTOTAL LOTE'):
                        for col_idx in range(1, len(datos_finales.columns) + 1):
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            cell.font = font_negrita
                
                # Ajustar ancho de columnas automáticamente
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 30)  # Máximo 30 caracteres
                    worksheet.column_dimensions[column_letter].width = adjusted_width

    print(f"\n✅ Proceso terminado. Archivo generado: {output_file}")
    print(f"📄 Se crearon las siguientes hojas por fecha:")
    
    # Mostrar hojas por fecha
    for fecha in sorted(fechas_unicas):
        cantidad = len(resumen[resumen['Fecha'] == fecha])
        # Formatear fecha de forma limpia para mostrar
        try:
            if hasattr(fecha, 'strftime'):
                fecha_display = fecha.strftime("%d-%m-%Y")
            else:
                fecha_display = str(fecha).replace('/', '-').replace(' 00:00:00', '')[:10]
        except:
            fecha_display = str(fecha)[:10]
        
        print(f"     • Fecha {fecha_display}: {cantidad} registros")
    
    # 📊 MOSTRAR ESTADÍSTICAS FINALES
    print(f"\n📊 Estadísticas finales:")
    print(f"     💰 Monto total procesado: ${resumen['Monto_Total'].sum():,.0f}")
    print(f"     🔢 Total de transacciones: {resumen['Cantidad_Transacciones'].sum()}")
    print(f"     🏪 Cajas procesadas: {len(resumen['Caja_Nombre'].unique())}")
    print(f"     💳 Tipos de tarjeta: {', '.join(sorted(resumen['Tarjeta'].unique()))}")

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"❌ Error inesperado: {e}")
        if os.name == 'nt':  # Windows
            input("Presiona Enter para cerrar...")
        import traceback
        traceback.print_exc()
